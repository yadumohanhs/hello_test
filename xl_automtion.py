import openpyxl as xl #xl is the alias name of openpyxl
from openpyxl.chart import BarChart, Reference #This imports Barchart from the openpyxl
wb=xl.load_workbook("transactions.xlsx")
sheet=wb['Sheet1']#for accessing the sheet 1 of spreadsheet
cell=sheet['a1']#for accessing the cells
cell=sheet.cell(1,1)#we can also use this format to acess cells.Here this means we are accessing row 1 column 1
print(cell.value)
#for finding the number of rows and printing it,we use "print(sheet.max_row)"
for row in range(2,sheet.max_row+1):  #Here we use +1 to include the number '4'.We also need to exclude the first row containing the text,'transaction_id'
    cell=sheet.cell(row,3)#This statement says for the values from the range(row) , the column will be the third column.this accesses the corresponding cells
    print(cell.value)
    corrected_price=cell.value*0.9#This decreases 10% of the original price
    corrected_price_cell=sheet.cell(row,4)#This creates a new column in the spread sheet
    corrected_price_cell.value=corrected_price

values=Reference(sheet,
                 min_row=2,
                 max_row=sheet.max_row,
                 min_col=2,
                 max_col=2
                 )
chart=BarChart()
chart.add_data(values)
sheet.add_chart(chart,'e2')

wb.save('transactions2.xlsx')    
